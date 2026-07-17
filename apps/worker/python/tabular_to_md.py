#!/usr/bin/env python3
"""确定性表格解析：csv/tsv/xlsx/xls → markdown 表格（逐行保真，无模型）。
用法：python tabular_to_md.py <文件路径>   → markdown 打到 stdout。
xlsx 每个非空 sheet 输出 `# {sheet名}` + 整表；csv 直接输出整表。"""
import sys
import os
import csv
import io


def cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).replace("\r", " ").replace("\n", " ").replace("|", "\\|").strip()


def _is_num(s):
    try:
        float(str(s).replace(",", "").replace("，", ""))
        return True
    except (ValueError, TypeError):
        return False


def table_md(rows):
    """rows: list[list]。第一行表头，其余数据行。列数按最大列对齐。
    丢全空行 + 丢全空列（修 Excel used-range 被撑到 16384 空列的情况）。"""
    rows = [r for r in rows if any(cell(c) != "" for c in r)]  # 丢全空行
    if not rows:
        return ""
    ncol = max(len(r) for r in rows)
    rows = [list(r) + [None] * (ncol - len(r)) for r in rows]
    # 丢「全空列」：任何一行在该列非空才保留（防 16384 空管道列撑爆 md）
    keep = [j for j in range(ncol) if any(cell(r[j]) != "" for r in rows)]
    if len(keep) < ncol:
        rows = [[r[j] for j in keep] for r in rows]
        ncol = len(keep)
    if ncol == 0:
        return ""
    # 丢开头「标题横幅行」：整行非空值全相同、且该值是「非数值文本」（多为公司名/表标题；排除 5,5,5 这种均匀数值 CSV）、
    # 下一行有 ≥2 个不同非空值（更像真实表头）。双护栏防误删数据：① 非数值 ② 删后必须仍剩 ≥1 数据行（len(rows)>=3）。
    while ncol >= 2 and len(rows) >= 3:
        cur = [cell(c) for c in rows[0] if cell(c) != ""]
        nxt = [cell(c) for c in rows[1] if cell(c) != ""]
        if cur and len(set(cur)) == 1 and not _is_num(cur[0]) and len(set(nxt)) >= 2:
            rows = rows[1:]
        else:
            break
    header = rows[0]
    out = []
    out.append("| " + " | ".join(cell(c) for c in header) + " |")
    out.append("| " + " | ".join(["---"] * ncol) + " |")
    for r in rows[1:]:
        out.append("| " + " | ".join(cell(c) for c in r) + " |")
    return "\n".join(out)


def _decode_bytes(raw):
    """CSV 编码嗅探：utf-8-sig(剥BOM,严格) → gb18030(GBK/GB2312超集,中文Windows CSV) → utf-8+replace 兜底不崩。
    已知限制：gb18030 对多数字节流都能无异常解码，故非中文 CJK(日文Shift-JIS/韩文EUC-KR/繁体Big5) 会被静默解成
    中文乱码而非报错——本工具面向中文场景，此为可接受的 minor 限制（用统计式 charset_normalizer 在短样本上反而
    会误判 GBK 中文、打破主用例，故不用）。"""
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    try:
        return raw.decode("gb18030")
    except UnicodeDecodeError:
        return raw.decode("utf-8", errors="replace")


def parse_csv(path):
    delim = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, "rb") as f:
        raw = f.read()
    rows = list(csv.reader(io.StringIO(_decode_bytes(raw)), delimiter=delim))
    return table_md(rows)


def _fill_merged(ws, grid):
    """把合并单元格左上角的值回填到整个合并区（否则行级切片会丢产品名/类别）。就地修改。"""
    for rng in list(ws.merged_cells.ranges):
        r1, r2, c1, c2 = rng.min_row, rng.max_row, rng.min_col, rng.max_col
        if r1 - 1 >= len(grid) or c1 - 1 >= len(grid[r1 - 1]):
            continue
        tl = grid[r1 - 1][c1 - 1]
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if r - 1 < len(grid) and c - 1 < len(grid[r - 1]):
                    grid[r - 1][c - 1] = tl


def parse_xlsx(path):
    import openpyxl

    # 双载：data_only=True 取 Excel 缓存的计算值；没有缓存(如 openpyxl 生成且未被 Excel 打开过)
    # 则回退到公式串，避免公式列变空白。真实 Excel 存的文件直接显示计算结果。
    wb_v = openpyxl.load_workbook(path, data_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    parts = []
    for ws_v, ws_f in zip(wb_v.worksheets, wb_f.worksheets):
        rows = []
        for rv, rf in zip(ws_v.iter_rows(values_only=True), ws_f.iter_rows(values_only=True)):
            rows.append([v if v is not None else f for v, f in zip(rv, rf)])
        _fill_merged(ws_v, rows)  # 合并单元格纵向/横向回填
        md = table_md(rows)
        if not md:
            continue
        parts.append(f"# {ws_v.title}\n\n{md}")
    return "\n\n".join(parts)


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("用法: python tabular_to_md.py <文件路径>\n")
        sys.exit(2)
    path = sys.argv[1]
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv"):
        md = parse_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        md = parse_xlsx(path)
    else:
        sys.stderr.write(f"不支持的表格类型: {ext}\n")
        sys.exit(2)
    sys.stdout.write(md.strip() + "\n")


if __name__ == "__main__":
    main()
